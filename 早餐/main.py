# region 导入函式库
import openpyxl
import reformat
import refund
from openpyxl import load_workbook
import os
import scratch
# endregion

datethedate = scratch.download_xls()
# datethedate = input('time')
# print('datethedate:', datethedate)
scratch.transfer_to_excel(scratch.find_path())
# region 导入数据
data_directory = os.path.dirname(os.path.abspath(__file__))
data_directory = data_directory.strip('\\早餐')
date = '9.19~9.25'

# --- 激活工作簿 -- #
target_file = data_directory + '/营收详情导出.xlsx'
wb1 = load_workbook(target_file)
ws1 = wb1.active

ex_file = data_directory + '/2022上學期早餐結算表.xlsx'
wb2 = load_workbook(ex_file)
ws2 = wb2.active
# ----------------- #


target_sheet = 'Sheet1'
change_u = wb1[target_sheet]  # 读取营收详情导出资料
target = wb2[date]  # 读取结算表表格
wb2.copy_worksheet(target)
be_change = wb2[date + ' Copy']  # 得到结算表复制表格
change_u.delete_rows(1, 1)
change_u.delete_cols(1, 1)
# endregion


wb1.save(target_file)
wb1.close()

count_shop = 1
while True:
    if change_u.cell(row=count_shop, column=1).value:
        count_shop = count_shop + 1
    else:
        break

count_food = 1
while True:
    if (be_change.cell(row=count_food, column=1).value is None) or (
            be_change.cell(row=count_food, column=1).value):
        if be_change.cell(row=count_food, column=1).value == '总计':
            break
        else:
            count_food = count_food + 1
count_food = count_food - 2

count_employee = 1
while True:
    if (be_change.cell(row=count_employee, column=1).value is None) or (
            be_change.cell(row=count_employee, column=1).value):
        if be_change.cell(row=count_employee, column=1).value == '珍德粤点 员工餐自购':
            break
        else:
            count_employee = count_employee + 1

print('营收详情：', count_shop)
print('员工餐：', count_employee)
print('food:', count_food)

change_u.delete_rows(count_shop - 1, count_shop - 1)
count_shop = count_shop - 1
print('营收详情：', count_shop)

for i in range(2, 100):
    if change_u.cell(row=i, column=1).value == '焙醇ripebakery':
        change_u.cell(row=i, column=1).value = 'ripebakery面包店'

for i in range(2, 100):
    if be_change.cell(row=i, column=1).value == '退款':
        be_change.cell(row=i, column=2).value = refund.refund(count_shop, change_u)
        print(be_change.cell(row=i, column=2).value)
change_u, be_change = reformat.shop_move_pos(count_employee, target, be_change, change_u)
profit_shop_list = reformat.read_shop_list_profit(data_directory)
change_u, be_change = refund.clean_profit(count_employee, profit_shop_list, change_u, be_change)  # 毛利润整理+清空毛利润
for r in range(2, count_employee+5):
    for c in range(2, 9):
        if change_u.cell(row=r, column=c).value:
            be_change.cell(row=r, column=c).value = round(eval(change_u.cell(row=r, column=c).value), 2)
# be_change.cell(row=7, column=8).value = be_change.cell(row=7, column=8).value + be_change.cell(row=6, column=8).value
# be_change.cell(row=6, column=8).value = None

# be_change.title = input('请输入日期：')
be_change.title = datethedate
wb2.save(ex_file)
wb2.close()



