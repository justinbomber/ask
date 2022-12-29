# region 导入函式库
import reformat
import refund
from openpyxl import load_workbook
import os
from formula import eval_all
import scratchu
# endregion



datethedate = scratchu.download_xls()
scratchu.transfer_to_excel(scratchu.find_path())

# region 导入数据
data_directory = os.path.dirname(os.path.abspath(__file__))
data_directory = data_directory.strip('\\new_ask')
date = '9.19~9.25'


# --- 激活工作簿 -- #
target_file = data_directory + '/ask/营收详情导出.xlsx'
wb1 = load_workbook(target_file)
ws1 = wb1.active

ex_file = data_directory + '/ask/2022上學期南村結算表.xlsx'
wb2 = load_workbook(ex_file)
ws2 = wb2.active
# ----------------- #


target_sheet = 'Sheet1'
change_u = wb1[target_sheet]  # 读取营收详情导出资料
target = wb2[date]  # 读取结算表表格
wb2.copy_worksheet(target)
be_change = wb2[date + ' Copy']  # 得到结算表复制表格
change_u.delete_rows(1, 1)
change_u.delete_cols(1, 1)
# endregion

# for r in range(2,10):
#     for c in range(1, 20):
#         if target.cell(row=r, column=c).value:
#             print(target.cell(row=r, column=c).value, ',type:', isinstance(target.cell(row=r, column=c).value, str))




count_shop = 1
while True:
    if change_u.cell(row=count_shop, column=1).value:
        count_shop = count_shop + 1
    else:
        break


count_food = 1
while True:
    if (be_change.cell(row=count_food, column=1).value is None) or (
            be_change.cell(row=count_food, column=1).value):
        if be_change.cell(row=count_food, column=1).value == '饮料分区':
            break
        else:
            count_food = count_food + 1
count_food = count_food - 2

count_you_down = 1
while True:
    if (be_change.cell(row=count_you_down, column=1).value is None) or (
            be_change.cell(row=count_you_down, column=1).value):
        if be_change.cell(row=count_you_down, column=1).value == '润心牛奶甜品':
            break
        else:
            count_you_down = count_you_down + 1

print('营收详情：', count_shop)
print('结算：', count_you_down)
print('food:', count_food)
change_u.delete_rows(count_shop-1, count_shop-1)
count_shop = count_shop - 1

change_u = refund.refund(count_shop, change_u)  # 退款转换完成 排版前
sum_s_define, uncle_se_define, g_s_define, sf_s_define= refund.s_define(change_u, count_shop)  # 自定义数据搜集完成
change_u = reformat.change_name(count_shop, change_u)  # 重命名
change_u, be_change = reformat.shop_move_pos(count_you_down, target, be_change, change_u, count_food)  # 排版完成+清空数据
change_u = refund.refund_combine(count_you_down, change_u)  # 退款合拼
change_u = eval_all(count_you_down, change_u)  # 数字化
change_u = refund.profit_combine(count_you_down, change_u)  # 利润合拼


profit_shop_list = reformat.read_shop_list_profit(data_directory)
change_u, be_change = refund.clean_profit(count_you_down, profit_shop_list, change_u, be_change)  # 毛利润整理+清空毛利润

# --- 输入档案 --- #
for r in range(2, 500):
    for c in range(2, 20):
        if change_u.cell(row=r, column=c).value:
            be_change.cell(row=r, column=c).value = change_u.cell(row=r, column=c).value


for r in range(65, 500):
    if be_change.cell(row=r, column=6).value:
        if be_change.cell(row=r, column=6).value == '阿叔':
            be_change.cell(row=r, column=7).value = uncle_se_define
        elif be_change.cell(row=r, column=6).value == '自定义费':
            be_change.cell(row=r, column=7).value = sum_s_define
        elif be_change.cell(row=r, column=6).value == '居肉町':
            be_change.cell(row=r, column=7).value = g_s_define
        elif be_change.cell(row=r, column=6).value == '丰顺捆粄':
            be_change.cell(row=r, column=7).value = sf_s_define

if be_change.cell(row=2, column=9).value:
    if be_change.cell(row=4, column=9).value:
        be_change.cell(row=4, column=9).value = be_change.cell(row=2, column=9).value + be_change.cell(row=4, column=9).value
        be_change.cell(row=2, column=9).value = None
    else:
        be_change.cell(row=4, column=9).value = be_change.cell(row=2, column=9).value
        be_change.cell(row=2, column=9).value = None


mc_money = 0
be_change.cell(row=3, column=7).value = mc_money
be_change.title = datethedate
# be_change.title = input('dddate')
wb2.save(ex_file)
wb2.close()



