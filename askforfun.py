# region 导入函式库
from openpyxl import load_workbook
import os
# endregion

# region 导入数据
data_directory = os.path.dirname(os.path.abspath(__file__))
date = '4.11~4.17'
target_file = data_directory + '/营收详情导出.xlsx'
target_sheet = 'Sheet1'
ex_file = data_directory + '/2022年下学期南村结算表.xlsx'
ex_sheet = date
wb1 = load_workbook(target_file)
ws1 = wb1.active
wb2 = load_workbook(ex_file)
ws2 = wb2.active
sample_sheet = wb1[target_sheet]  # input
example_sheet = wb2[ex_sheet]  # register
wb2.copy_worksheet(example_sheet)
final_sheet = wb2[date + ' Copy']  # final_output
sample_sheet.delete_rows(1, 1)
# endregion

# region 清空毛利润
for i in range(1, 150):
    if final_sheet.cell(row=i, column=12).value:
        final_sheet.cell(row=i, column=12).value = None
# endregion

# region换名字
for i in range(2, 100):
    if (sample_sheet.cell(row=i, column=1).value):
        if sample_sheet.cell(row=i, column=1).value == '芝士王茶•轻食&饮品(T1~5)':
            sample_sheet.cell(row=i, column=1).value = 'Mr.king芝士王茶(T1~5)'
        if sample_sheet.cell(row=i, column=1).value == '芝士王茶•轻食&饮品(T10~12)':
            sample_sheet.cell(row=i, column=1).value = 'Mr.king芝士王茶(T10~12)'
        if sample_sheet.cell(row=i, column=1).value == '彩虹绵绵冰':
            sample_sheet.cell(row=i, column=1).value = '杨小贤 芒果绵绵冰'
        if sample_sheet.cell(row=i, column=1).value == '遇见小面 活动 (T1-5)':
            sample_sheet.cell(row=i, column=1).value = '遇见小面 (T1-5)'
        if sample_sheet.cell(row=i, column=1).value == '遇见小面 活动 (T10-12)':
            sample_sheet.cell(row=i, column=1).value = '遇见小面 (T10-12)'

# endregion

# region 计数
Count_shop_new = 1
while True:
    if sample_sheet.cell(row=Count_shop_new, column=1).value:
        Count_shop_new = Count_shop_new + 1
    else:
        break

# endregion

# region 退款计算
recall = 0
for i in range(2, Count_shop_new):
    for j in range(len(sample_sheet.cell(row=i, column=8).value)):
        if sample_sheet.cell(row=i, column=8).value[j] == '：':
            a = sample_sheet.cell(row=i, column=8).value[j + 1:]
            sample_sheet.cell(row=i, column=8).value = a
            if a:
                a = a[1:]
                recall = recall + float(a)
                sample_sheet.cell(row=i, column=8).value = round(float(a), 2)
                break
# endregion

# region 自定义费用
selfdefine = 0

for i in range(70, 90):
    if final_sheet.cell(row=i, column=6).value == '阿叔':
        final_sheet.cell(row=i, column=7).value = 0
for i in range(70, 90):
    if final_sheet.cell(row=i, column=6).value == '北飘':
        final_sheet.cell(row=i, column=7).value = 0
for r in range(2, 70):
    if sample_sheet.cell(row=r, column=1).value == '阿叔猪扒包':
        for i in range(70, 90):
            if final_sheet.cell(row=i, column=6).value == '阿叔':
                final_sheet.cell(row=i, column=7).value = round(eval(sample_sheet.cell(row=r, column=11).value), 1)
    elif sample_sheet.cell(row=r, column=1).value == '北飘久香 烧烤':
        for i in range(70, 90):
            if final_sheet.cell(row=i, column=6).value == '北飘':
                final_sheet.cell(row=i, column=7).value = round(eval(sample_sheet.cell(row=r, column=11).value), 1)
for r in range(2, 70):
    if sample_sheet.cell(row=r, column=11).value:
        selfdefine = selfdefine + round(eval(sample_sheet.cell(row=r, column=11).value), 1)
for r in range(70, 90):
    if final_sheet.cell(row=r, column=6).value == '自定义费':
        final_sheet.cell(row=r, column=7).value = selfdefine
        break
# endregion

# region 排版
sample_sheet.move_range('F1:F100', 0, 17, True)
sample_sheet.move_range('P1:P100', 0, 6, True)
sample_sheet.move_range('I1:I100', 0, 16, True)
sample_sheet.move_range('Q1:Q100', 0, 15, True)
sample_sheet.move_range('H1:H100', 0, 21, True)
sample_sheet.delete_cols(2, 20)

# endregion

# region 数据字典
list_col = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U',
            'V', 'W', 'X', 'Y', 'Z']
# endregion


# region 创建商家字典
dict_shop = {}
countyoudown = 1
while True:
    if (final_sheet.cell(row=countyoudown, column=1).value == None) or (
    final_sheet.cell(row=countyoudown, column=1).value):
        if final_sheet.cell(row=countyoudown, column=1).value == '本该有咖啡':
            break
        else:
            countyoudown = countyoudown + 1

for i in range(2, countyoudown + 1):
    if (example_sheet.cell(row=i, column=1).value and example_sheet.cell(row=i, column=1).value != '饮料分区'):
        dict_shop[example_sheet.cell(row=i, column=1).value] = i
    for j in range(2, 6):
        final_sheet.cell(row=i, column=j).value = None
# endregion

# region 毛利润计算
for i in range(2, countyoudown + 1):
    if sample_sheet.cell(row=i, column=12).value:
        sample_sheet.cell(row=i, column=12).value = round(eval(sample_sheet.cell(row=i, column=12).value), 2)
        x = sample_sheet.cell(row=i, column=1).value
        if x == '【御食家】·盖码饭' or x == '沪上阿姨':
            sample_sheet.cell(row=i, column=12).value = None
            continue
        elif sample_sheet.cell(row=i, column=1).value[0] == sample_sheet.cell(row=i - 1, column=1).value[0]:
            sample_sheet.cell(row=i, column=12).value = sample_sheet.cell(row=i, column=12).value + sample_sheet.cell(
                row=i - 1, column=12).value
            sample_sheet.cell(row=i - 1, column=12).value = None
# endregion

# region 移位(insertion sort)
for i in range(2, Count_shop_new):
    s = 'A' + str(i) + ':AZ' + str(i)
    sample_sheet.move_range(s, rows=300, cols=0, translate=True)
counts = 0
for i in range(302, Count_shop_new + 300):
    s = 'A' + str(i) + ':AZ' + str(i)
    if dict_shop[sample_sheet.cell(row=i, column=1).value]:
        sample_sheet.move_range(s, rows=dict_shop[sample_sheet.cell(row=i, column=1).value] - i, cols=0, translate=True)

# endregion

# region 数据导入
for r in range(2, 100):
    final_sheet.cell(row=r, column=9).value = None
    final_sheet.cell(row=r, column=11).value = None
for r in range(2, countyoudown + 1):
    for c in range(2, 6):
        if sample_sheet.cell(row=r, column=c).value:
            sample_sheet.cell(row=r, column=c).value = eval(sample_sheet.cell(row=r, column=c).value)
for r in range(2, countyoudown + 1):
    for c in range(2, 13):
        if sample_sheet.cell(row=r, column=c).value:
            final_sheet.cell(row=r, column=c).value = sample_sheet.cell(row=r, column=c).value
final_sheet.cell(row=4, column=7).value = 0
for r in range(2, 150):
    if final_sheet.cell(row=r, column=11).value:
        final_sheet.cell(row=r, column=11).value = None

# endregion

# region 退款合拼
for r in range(2, 65):
    if final_sheet.cell(row=r, column=1).value and final_sheet.cell(row=r + 1, column=1).value:
        if final_sheet.cell(row=r, column=1).value[0] == final_sheet.cell(row=r + 1, column=1).value[0]:
            if final_sheet.cell(row=r, column=9).value == None and final_sheet.cell(row=r + 1, column=9).value == None:
                continue
            elif final_sheet.cell(row=r, column=9).value == None and final_sheet.cell(row=r + 1, column=9).value:
                continue
            elif final_sheet.cell(row=r, column=9).value and final_sheet.cell(row=r + 1, column=9).value == None:
                final_sheet.cell(row=r, column=9).value, final_sheet.cell(row=r + 1, column=9).value = final_sheet.cell(
                    row=r + 1, column=9).value, final_sheet.cell(row=r, column=9).value
            elif final_sheet.cell(row=r, column=9).value and final_sheet.cell(row=r + 1, column=9).value:
                final_sheet.cell(row=r + 1, column=9).value = final_sheet.cell(row=r,
                                                                               column=9).value + final_sheet.cell(
                    row=r + 1, column=9).value
                final_sheet.cell(row=r, column=9).value = None
# final_sheet.cell(row=4,column=9).value = final_sheet.cell(row=4,column=9).value + final_sheet.cell(row=2,
# column=9).value final_sheet.cell(row=2,column=9).value = None endregion

# region 公式载入
formula_shop_set = {'三叔粥铺(T10~12)', '杨小贤 芒果绵绵冰', '益禾堂T10~12', '书亦烧仙草'}
for r in range(2, countyoudown + 1):
    if example_sheet.cell(row=r, column=1).value:
        if example_sheet.cell(row=r, column=1).value in formula_shop_set:
            final_sheet.cell(row=r, column=12).value = example_sheet.cell(row=r, column=12).value
final_sheet.cell(row=40, column=12).value = '=F40*0.221'
# endregion


a = input('请输入日期：')
final_sheet.title = a
wb2.save(ex_file)
wb2.close()
